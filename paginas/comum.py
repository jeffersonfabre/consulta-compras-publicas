"""
Helpers compartilhados entre as páginas do app.

Inclui:
- Formatação (_to_float, _formatar_cnpj, _formatar_brl, _formatar_data)
- Wrappers com cache (@st.cache_data) das funções do api_client
- Gerador de XLSX genérico (cabeçalho + tabela + linha de total + aba de detalhes)
"""
from __future__ import annotations

import io
from typing import Any

import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import api_client

# =============================================================================
# Formatação
# =============================================================================

def to_float(v) -> float:
    """Converte qualquer coisa para float, tratando None e strings."""
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def formatar_cnpj(cnpj: str) -> str:
    """Formata CNPJ no padrão 00.000.000/0000-00."""
    if not cnpj:
        return ""
    apenas_num = "".join(c for c in str(cnpj) if c.isdigit())
    if len(apenas_num) != 14:
        return str(cnpj)
    return f"{apenas_num[:2]}.{apenas_num[2:5]}.{apenas_num[5:8]}/{apenas_num[8:12]}-{apenas_num[12:]}"


def formatar_brl(valor: float) -> str:
    """Formata um número como moeda BRL: R$ 1.234,56"""
    return f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def formatar_data(d: str) -> str:
    """Devolve só a parte YYYY-MM-DD de uma data ISO, ou string vazia."""
    return (d or "")[:10]


# =============================================================================
# Wrappers com cache (TTL = 1h, configurável)
# =============================================================================
# A API é pública e gratuita; cachear evita refazer as mesmas chamadas várias
# vezes na mesma sessão. show_spinner=False porque já temos progress próprio.

CACHE_TTL = 3600  # 1 hora


@st.cache_data(ttl=CACHE_TTL, show_spinner=False)
def cache_buscar_contratacoes(uasg: str, codigo_modalidade: int, ano: int) -> list[dict]:
    return api_client.buscar_contratacoes(uasg, codigo_modalidade, ano)


@st.cache_data(ttl=CACHE_TTL, show_spinner=False)
def cache_buscar_itens(uasg: str, ano: int, orgao_cnpj: str | None = None) -> list[dict]:
    return api_client.buscar_itens(uasg, ano, orgao_cnpj=orgao_cnpj)


@st.cache_data(ttl=CACHE_TTL, show_spinner=False)
def cache_buscar_resultados(uasg: str, ano: int, orgao_cnpj: str | None = None) -> list[dict]:
    return api_client.buscar_resultados(uasg, ano, orgao_cnpj=orgao_cnpj)


@st.cache_data(ttl=CACHE_TTL, show_spinner=False)
def cache_buscar_arps(
    uasg_gerenciadora: str, ano: int, codigo_modalidade: str | None = None
) -> list[dict]:
    return api_client.buscar_arps(uasg_gerenciadora, ano, codigo_modalidade=codigo_modalidade)


@st.cache_data(ttl=CACHE_TTL, show_spinner=False)
def cache_buscar_arp_itens(
    uasg_gerenciadora: str, ano: int, codigo_modalidade: str | None = None
) -> list[dict]:
    return api_client.buscar_arp_itens(uasg_gerenciadora, ano, codigo_modalidade=codigo_modalidade)


@st.cache_data(ttl=CACHE_TTL, show_spinner=False)
def cache_buscar_contratos(
    uasg_gestora: str, ano: int, numero_contrato: str | None = None
) -> list[dict]:
    return api_client.buscar_contratos(uasg_gestora, ano, numero_contrato=numero_contrato)


@st.cache_data(ttl=CACHE_TTL, show_spinner=False)
def cache_buscar_itens_contrato(
    uasg_gestora: str, ano: int, numero_contrato: str | None = None
) -> list[dict]:
    return api_client.buscar_itens_contrato(uasg_gestora, ano, numero_contrato=numero_contrato)


# =============================================================================
# Gerador genérico de XLSX
# =============================================================================

def gerar_xlsx(
    df: pd.DataFrame,
    titulo: str,
    subtitulo: str,
    descricao: str,
    detalhes: list[tuple[str, Any]],
    nome_aba_detalhes: str = "Detalhes",
    coluna_valor_total: str = "Valor Total (R$)",
    colunas_monetarias: tuple[str, ...] = (),
    colunas_quantidade: tuple[str, ...] = (),
    larguras: dict[str, int] | None = None,
) -> bytes:
    """
    Gera um xlsx formatado com:
      - Cabeçalho com título + subtítulo + descrição (linhas 1-3)
      - Tabela na linha 5 com cabeçalho preto/branco
      - Formatação monetária e numérica nas colunas indicadas
      - Linha TOTAL no final somando `coluna_valor_total` (se existir)
      - Aba secundária com detalhes (lista de tuplas campo/valor)
    """
    output = io.BytesIO()
    larguras = larguras or {}

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # === Aba principal ===
        df.to_excel(writer, sheet_name="Itens", index=False, startrow=4)
        ws = writer.sheets["Itens"]

        # Cabeçalho informativo
        ws["A1"] = titulo
        ws["A1"].font = Font(bold=True, size=14, color="0F172A")
        ws["A2"] = subtitulo
        ws["A2"].font = Font(size=11, color="64748B")
        ws["A3"] = descricao
        ws["A3"].font = Font(size=10, color="64748B")

        # Estiliza cabeçalho da tabela (linha 5)
        header_fill = PatternFill("solid", start_color="0F172A")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=5, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Larguras das colunas
        for idx, col in enumerate(df.columns, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = larguras.get(col, 18)

        # Formatos numéricos
        n_linhas_dados = len(df) + 5
        for row in range(6, n_linhas_dados + 1):
            for col_name in colunas_monetarias:
                if col_name in df.columns:
                    col_idx = list(df.columns).index(col_name) + 1
                    ws.cell(row=row, column=col_idx).number_format = (
                        '"R$" #,##0.00;[Red]"R$" -#,##0.00'
                    )
            for col_name in colunas_quantidade:
                if col_name in df.columns:
                    col_idx = list(df.columns).index(col_name) + 1
                    ws.cell(row=row, column=col_idx).number_format = "#,##0.##"

        # Linha de total (se a coluna existir e tiver dados)
        if not df.empty and coluna_valor_total in df.columns:
            total_row = n_linhas_dados + 1
            ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
            valor_total_col = list(df.columns).index(coluna_valor_total) + 1
            cell_total = ws.cell(
                row=total_row,
                column=valor_total_col,
                value=f"=SUM({get_column_letter(valor_total_col)}6:"
                f"{get_column_letter(valor_total_col)}{n_linhas_dados})",
            )
            cell_total.font = Font(bold=True, color="0F172A")
            cell_total.number_format = '"R$" #,##0.00'
            cell_total.fill = PatternFill("solid", start_color="F0FDF4")

        # === Aba secundária: detalhes ===
        df_detalhes = pd.DataFrame(detalhes, columns=["Campo", "Valor"])
        df_detalhes.to_excel(writer, sheet_name=nome_aba_detalhes, index=False)
        ws2 = writer.sheets[nome_aba_detalhes]
        ws2.column_dimensions["A"].width = 28
        ws2.column_dimensions["B"].width = 60
        for col_idx in (1, 2):
            cell = ws2.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font

    output.seek(0)
    return output.getvalue()
